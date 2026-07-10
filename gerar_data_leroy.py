#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Gera o data.json do dashboard Leroy Merlin (status-leroy-maneng)
a partir da planilha Google "Leroy Merlin Base de Dados" (export xlsx).

Uso: python3 gerar_data_leroy.py <planilha.xlsx> <data_json_anterior> <saida.json> [--gerado YYYY-MM-DD]

Preserva do data.json anterior: corretivas, preventivas e a ordem/identidade
do ranking de criticidade (fonte: Indicação Maneng). Recalcula todo o resto.
"""
import sys, json, unicodedata, datetime
import openpyxl

ABAS_NAO_LOJA = {'EFICIÊNCIA', 'EFICIÊNCIA REGIÕES', 'STATUS LOJA', 'PLANO DE AÇÃO'}
LITORAL_SEQ = ['MACEIO', 'NATAL', 'FORTALEZA', 'RIO BARRA', 'RIO NORTE', 'NITEROI']
LITORAL_FLAG = {'MACEIO', 'NATAL', 'FORTALEZA', 'RIO BARRA', 'NITEROI'}

def norm(s):
    s = unicodedata.normalize('NFD', str(s or ''))
    return ''.join(c for c in s if unicodedata.category(c) != 'Mn').upper().strip()

def cell_str(v):
    if v is None: return ''
    if isinstance(v, float) and v == int(v): return str(int(v))
    if isinstance(v, datetime.datetime): return v.strftime('%d/%m/%Y')
    return str(v).strip()

def is_salao(a):
    n = norm(a); return 'SALAO' in n and 'VENDA' in n

def tipo_equip(e):
    n = norm(e)
    if 'FANCOIL' in n: return 'Fancoil'
    if 'CHILLER' in n: return 'Chiller'
    if 'SPLIT TOP' in n or 'SPLITTOP' in n: return 'Split Top'
    if 'SPLITAO' in n: return 'Splitão'
    if 'ROOF' in n: return 'Roof Top'
    if 'PISO TETO' in n: return 'Piso Teto'
    return 'Não especificado'

def parse_store(ws, prev_meta):
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    title, hdr_i, hdr = '', None, None
    for i, r in enumerate(rows[:8]):
        vals = [cell_str(c) for c in r]
        if not title:
            for v in vals:
                if v and 'LEROY' in v.upper(): title = v; break
        if any(norm(c) == 'AMBIENTE' for c in vals):
            hdr_i, hdr = i, vals; break
    if hdr_i is None: return None
    def col(*names):
        for j, h in enumerate(hdr):
            if norm(h) in [norm(n) for n in names]: return j
        return None
    c_amb, c_eq = col('AMBIENTE'), col('EQUIPAMENTO')
    c_ev, c_co = col('MODELO EVAP.', 'MODELO EVAP'), col('MODELO COND.', 'MODELO COND')
    c_f = col('FABRICAÇÃO', 'FABRICACAO')
    c_p, c_o = col('PENDÊNCIA', 'PENDENCIA'), col('OBSERVAÇÕES', 'OBSERVACOES', 'OBSERVAÇÃO')
    circ_cols = [(j, hdr[j].strip()) for j in range(len(hdr))
                 if len(hdr[j].strip()) == 1 and hdr[j].strip().isalpha()]
    equip = []
    for r in rows[hdr_i + 1:]:
        vals = [cell_str(c) if not (isinstance(c, str) and c.startswith('=')) else '' for c in r]
        def g(j): return vals[j] if j is not None and j < len(vals) else ''
        e_nome = g(c_eq)
        if norm(g(c_amb)) == 'SISTEMAS' or norm(e_nome) == 'QUANT': break  # bloco resumo no rodapé
        if not e_nome:
            if not any(g(j) != '' for j, _ in circ_cols): continue
            e_nome = 'Equipamento %02d' % (len(equip) + 1)
            sem_nome = True
        else:
            sem_nome = False
        circs = []
        for j, letra in circ_cols:
            v = g(j)
            if v != '': circs.append({'l': letra, 'k': 1 if v == '1' else 0})
        if not circs: circs = [{'l': 'A', 'k': 1}]
        ev, co, f = g(c_ev), g(c_co), g(c_f)
        cad = 1 if sem_nome else 0
        equip.append({'a': g(c_amb), 'e': e_nome, 'ev': ev, 'co': co, 'f': f,
                      'p': g(c_p), 'o': g(c_o), 'cad': cad, 'c': circs})
    meta = prev_meta or {}
    st = {'sheet': ws.title, 'title': title or ('LEROY MERLIN ' + ws.title),
          'loja': meta.get('loja', ''), 'spv': meta.get('spv', ''), 'uf': meta.get('uf', ''),
          'cadFalta': sum(e['cad'] for e in equip), 'equip': equip}
    return st

def parse_plano(wb):
    if 'PLANO DE AÇÃO' not in wb.sheetnames: return []
    rows = [list(r) for r in wb['PLANO DE AÇÃO'].iter_rows(values_only=True)]
    hdr_i = next((i for i, r in enumerate(rows) if any('REUNI' in norm(c) for c in r if c)), None)
    if hdr_i is None: return []
    out = []
    for r in rows[hdr_i + 1:]:
        v = [cell_str(c) for c in r]
        v += [''] * (10 - len(v))
        if not v[2] and not v[3]: continue
        out.append({'reuniao': v[1], 'loja': v[2], 'problema': v[3], 'data': v[4],
                    'maneng': v[5], 'leroy': v[6], 'diel': v[7], 'relatorio': v[8]})
    return out

def maquina_parada(e): return any(not c['k'] for c in e['c'])

def build(xlsx_path, prev_path, out_path, gerado=None):
    prev = json.load(open(prev_path, encoding='utf-8'))
    prev_meta = {s['sheet']: s for s in prev.get('stores', [])}
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    gerado = gerado or datetime.date.today().isoformat()
    ano_ref = int(gerado[:4])

    stores = []
    for name in wb.sheetnames:
        if name in ABAS_NAO_LOJA: continue
        st = parse_store(wb[name], prev_meta.get(name))
        if st: stores.append(st)

    # uf agregado: aba EFICIÊNCIA (UF, Total, Em Operação, Inoperante)
    uf = []
    if 'EFICIÊNCIA' in wb.sheetnames:
        rows_ef = [list(r) for r in wb['EFICIÊNCIA'].iter_rows(values_only=True)]
        hdr_i = next((i for i, r in enumerate(rows_ef)
                      if any(norm(c) == 'UF' for c in r if c)), None)
        if hdr_i is not None:
            hv = [norm(c) for c in rows_ef[hdr_i]]
            j_uf, j_t = hv.index('UF'), hv.index('TOTAL')
            j_o = next(j for j, h in enumerate(hv) if 'OPERA' in h)
            j_i = next(j for j, h in enumerate(hv) if 'INOPER' in h)
            for r in rows_ef[hdr_i + 1:]:
                v = [cell_str(c) for c in r] + [''] * 8
                if not v[j_uf] or len(v[j_uf]) > 4 or not v[j_t].isdigit(): continue
                uf.append({'uf': v[j_uf], 'total': int(v[j_t]),
                           'oper': int(v[j_o] or 0), 'inop': int(v[j_i] or 0)})

    # pareto (máquinas paradas, todas as áreas)
    par = [{'loja': s['sheet'], 'sheet': s['sheet'],
            'parados': sum(1 for e in s['equip'] if maquina_parada(e)), 'uf': s['uf']}
           for s in stores]
    par = sorted([p for p in par if p['parados'] > 0], key=lambda x: -x['parados'])
    total_par = sum(p['parados'] for p in par)
    ac = 0
    for p in par:
        ac += p['parados']
        p['acumPct'] = round(100 * ac / total_par, 1) if total_par else 0
    pareto = {'lojas': par, 'totalParados': total_par}
    top5 = sum(p['parados'] for p in par[:5])
    concentracao = {'top5': top5, 'total': total_par,
                    'pct': round(100 * top5 / total_par) if total_par else 0,
                    'top5lojas': [p['sheet'] for p in par[:5]]}

    # salão por tipo
    parados_tipo, totais_tipo = {}, {}
    for s in stores:
        for e in s['equip']:
            if not is_salao(e['a']): continue
            t = tipo_equip(e['e'])
            totais_tipo[t] = totais_tipo.get(t, 0) + 1
            if maquina_parada(e):
                parados_tipo.setdefault(t, []).append(
                    {'loja': s['sheet'], 'sheet': s['sheet'], 'eq': e['e'],
                     'pend': e['p'], 'circ': sum(1 for c in e['c'] if not c['k'])})
    salaoPorTipo = {'parados': parados_tipo, 'totais': totais_tipo}

    # obsolescência (salão) — todas as lojas
    obs_lojas = []
    for s in stores:
        maq_salao = [e for e in s['equip'] if is_salao(e['a'])]
        litoral = norm(s['sheet']) in LITORAL_FLAG
        limite = 5 if litoral else 10
        maqs, anos = [], []
        import re
        for e in maq_salao:
            m = re.search(r'(19|20)\d{2}', e['f'])
            ano = int(m.group(0)) if m else None
            idade = (ano_ref - ano) if ano is not None else None
            maqs.append({'eq': e['e'], 'ano': ano, 'idade': idade, 'evap': e['ev'],
                         'cond': e['co'], 'trocar': ano is not None and idade >= limite,
                         'limite': limite})
            if ano is not None: anos.append(ano)
        obs_lojas.append({'sheet': s['sheet'], 'title': s['title'], 'uf': s['uf'],
                          'maqSalao': len(maq_salao), 'comAno': len(anos),
                          'semAno': len(maq_salao) - len(anos), 'limiteTroca': limite,
                          'min': min(anos) if anos else None, 'max': max(anos) if anos else None,
                          'mista': bool(anos) and min(anos) != max(anos), 'anos': sorted(set(anos)),
                          'trocar': sum(1 for m in maqs if m['trocar']),
                          'litoral': litoral, 'maquinas': maqs})
    def obs_key(l):
        grp = 0 if (l['trocar'] > 0 and l['litoral']) else (1 if l['trocar'] > 0 else 2)
        return (grp, l['min'] if l['min'] is not None else 9999, -l['trocar'], l['sheet'])
    obs_lojas.sort(key=obs_key)
    for i, l in enumerate(obs_lojas): l['rank'] = i + 1
    obsolescencia = {'gerado': gerado, 'anoRef': ano_ref, 'lojas': obs_lojas, 'litoralSeq': LITORAL_SEQ}

    # criticidade: ranking manual preservado, métricas recalculadas
    crit_prev = prev.get('criticidade', {})
    by_sheet = {s['sheet']: s for s in stores}
    obs_by_sheet = {o['sheet']: o for o in obs_lojas}
    crit_lojas = []
    for c in crit_prev.get('lojas', []):
        s = by_sheet.get(c['sheet']); o = obs_by_sheet.get(c['sheet'])
        nc = dict(c)
        if s:
            tot = sum(len(e['c']) for e in s['equip'])
            oper = sum(1 for e in s['equip'] for x in e['c'] if x['k'])
            nc['pct'] = round(100 * oper / tot) if tot else 100
            parados = [e for e in s['equip'] if maquina_parada(e)]
            nc['parados'] = len(parados)
            nc['paradosLista'] = [{'amb': e['a'], 'eq': e['e'], 'pend': e['p'],
                                   'circ': sum(1 for x in e['c'] if not x['k'])} for e in parados]
            nc['title'] = s['title']; nc['uf'] = s['uf']; nc['spv'] = s['spv']
        if o:
            nc['trocar'] = o['trocar']
            nc['maqTrocar'] = [m for m in o['maquinas'] if m['trocar']]
            nc['litoral'] = o['litoral']; nc['anos'] = o['anos']
        crit_lojas.append(nc)
    criticidade = {'gerado': gerado, 'fonte': crit_prev.get('fonte', 'Indicação Maneng Mantenedora'),
                   'lojas': crit_lojas}

    db = {'gerado': gerado, 'stores': stores, 'uf': uf, 'plano': parse_plano(wb),
          'corretivas': prev.get('corretivas', []), 'preventivas': prev.get('preventivas', []),
          'obsolescencia': obsolescencia, 'salaoPorTipo': salaoPorTipo,
          'paretoLojas': pareto, 'concentracao': concentracao, 'criticidade': criticidade}
    json.dump(db, open(out_path, 'w', encoding='utf-8'), ensure_ascii=False, separators=(',', ':'))
    print(f"OK: {len(stores)} lojas, {total_par} máquinas paradas, gerado={gerado} -> {out_path}")
    return db

if __name__ == '__main__':
    ger = None
    args = [a for a in sys.argv[1:] if not a.startswith('--gerado')]
    for a in sys.argv[1:]:
            if a.startswith("--gerado="): ger = a.split("=", 1)[1]
    build(args[0], args[1], args[2], ger)
